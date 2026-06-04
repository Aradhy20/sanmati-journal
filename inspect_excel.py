import openpyxl
import json

wb = openpyxl.load_workbook('papers.xlsx')
print("Sheets:", wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- {sheetname} ---")
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) > 0:
        print("Header:", rows[0])
    if len(rows) > 1:
        print("Row 1:", rows[1])
    if len(rows) > 2:
        print("Row 2:", rows[2])
    print("Total rows:", len(rows))
