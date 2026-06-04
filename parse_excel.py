import openpyxl
import json

wb = openpyxl.load_workbook('papers.xlsx')

data = {}
total_papers = 0

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    
    # Header is rows[0]
    headers = [h.strip() if h else "" for h in rows[0]]
    
    sheet_papers = []
    for r in rows[1:]:
        # Filter rows where all cells are empty
        if all(cell is None for cell in r):
            continue
        # Construct paper dict
        paper = {}
        for i, val in enumerate(r):
            if i < len(headers):
                h = headers[i]
                if h:
                    paper[h] = val.strip() if isinstance(val, str) else val
        if paper.get('Title') or paper.get('Title '):
            sheet_papers.append(paper)
            
    data[sheetname] = sheet_papers
    total_papers += len(sheet_papers)
    print(f"Sheet '{sheetname}': {len(sheet_papers)} papers found.")

print(f"Total papers found: {total_papers}")

with open('papers.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print("Saved to papers.json")
